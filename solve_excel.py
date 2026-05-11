import openpyxl
import math
import random
import subprocess
import sys

# Ensure scipy is installed for stats distribution
try:
    import scipy.stats as stats
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "scipy"])
    import scipy.stats as stats

# Archivo de entrada y salida
input_file = r"D:\Estadistica Inferencial\Actividad de seleccion de muestra.xlsx"
output_file = r"D:\Estadistica Inferencial\Actividad de seleccion de muestra_Resuelto.xlsx"

# Cargar el libro de trabajo
wb = openpyxl.load_workbook(input_file)

# 1. Leer MATRIZ para obtener N y los datos
sheet_matriz = wb["MATRIZ"]
data_matriz = []
header = None
for i, row in enumerate(sheet_matriz.iter_rows(values_only=True)):
    if i == 0:
        header = row
    else:
        # Assuming first column is Id, stop if we hit empty rows
        if row[0] is not None:
            data_matriz.append(row)

N = len(data_matriz)

# 2. Cálculos Estadísticos
# Pregunta 1
confianza_1 = 0.92
alpha_1 = 1 - confianza_1
z_1 = stats.norm.ppf(1 - alpha_1/2) # 0.96 -> ~1.75
var_1 = 122500
e_1 = 100

n1_num = N * (z_1**2) * var_1
n1_den = (N - 1)*(e_1**2) + (z_1**2)*var_1
n1 = math.ceil(n1_num / n1_den)

# Pregunta 2
confianza_2 = 0.90
alpha_2 = 1 - confianza_2
z_2 = stats.norm.ppf(1 - alpha_2/2) # 0.95 -> ~1.645
p = 0.85
q = 0.15
e_2 = 0.08

n2_num = N * (z_2**2) * p * q
n2_den = (N - 1)*(e_2**2) + (z_2**2)*p * q
n2 = math.ceil(n2_num / n2_den)

# Elegir el mayor para la muestra común
n_final = max(n1, n2)

# 3. Escribir resultados en Actividad
sheet_actividad = wb["Actividad"]
# Buscamos un lugar libre al final
max_row = sheet_actividad.max_row
r = max_row + 2
sheet_actividad.cell(row=r, column=1, value="--- RESPUESTAS CALCULADAS AUTOMÁTICAMENTE ---")
sheet_actividad.cell(row=r+1, column=1, value=f"N (Total Población) obtenida de MATRIZ = {N} empresas.")

sheet_actividad.cell(row=r+3, column=1, value="Respuesta 1: Tamaño de muestra promedio")
sheet_actividad.cell(row=r+4, column=1, value=f"Z(92%) = {z_1:.4f}, Varianza = {var_1}, Error = {e_1}")
sheet_actividad.cell(row=r+5, column=1, value=f"Fórmula n1 calculada = {n1_num/n1_den:.2f}")
sheet_actividad.cell(row=r+6, column=1, value=f"-> Tamaño redondeado (n1) = {n1}")

sheet_actividad.cell(row=r+8, column=1, value="Respuesta 2: Tamaño de muestra proporción")
sheet_actividad.cell(row=r+9, column=1, value=f"Z(90%) = {z_2:.4f}, P = {p}, Q = {q}, Error = {e_2}")
sheet_actividad.cell(row=r+10, column=1, value=f"Fórmula n2 calculada = {n2_num/n2_den:.2f}")
sheet_actividad.cell(row=r+11, column=1, value=f"-> Tamaño redondeado (n2) = {n2}")

sheet_actividad.cell(row=r+13, column=1, value="Tamaño de muestra a utilizar para extraer (El mayor entre n1 y n2)")
sheet_actividad.cell(row=r+14, column=1, value=f"-> Tamaño final (n) = {n_final}")

# 4. Muestreo MAS
sheet_mas = wb["MAS"]
# Insert Headers
for col_idx, h in enumerate(header):
    sheet_mas.cell(row=1, column=col_idx+1, value=h)
sheet_mas.cell(row=1, column=len(header)+1, value="Aleatorio")

# Asignar aleatorios y ordenar
mas_data = []
for row in data_matriz:
    # Añadimos valor aleatorio
    mas_data.append((row, random.random()))

# Ordenamos por la columna aleatoria y cogemos los primeros n_final
mas_data.sort(key=lambda x: x[1])
mas_sample = mas_data[:n_final]

# Escribir en hoja MAS
for r_idx, (row_data, rand_val) in enumerate(mas_sample, start=2):
    for c_idx, val in enumerate(row_data):
        sheet_mas.cell(row=r_idx, column=c_idx+1, value=val)
    sheet_mas.cell(row=r_idx, column=len(header)+1, value=rand_val)

# 5. Muestreo Sistemático (MS)
sheet_ms = wb["MS"]
for col_idx, h in enumerate(header):
    sheet_ms.cell(row=1, column=col_idx+1, value=h)

# Calcular Salto y Arranque
K = math.floor(N / n_final)
if K < 1: K = 1 # Safety check
arranque = random.randint(1, K)

sheet_ms.cell(row=1, column=len(header)+2, value=f"Salto (K) = {K}")
sheet_ms.cell(row=2, column=len(header)+2, value=f"Arranque aleatorio = {arranque}")

# Seleccionar MS
ms_sample = []
current_index = arranque - 1 # Zero indexed array
while len(ms_sample) < n_final and current_index < N:
    ms_sample.append(data_matriz[current_index])
    current_index += K
    
# Si por truncamiento faltan datos, dar vuelta al principio (circular)
# Usualmente MS se asume estricto pero para llegar a 'n' sin problemas
if len(ms_sample) < n_final:
    current_index = (arranque - 1 + K*len(ms_sample)) % N
    while len(ms_sample) < n_final:
        ms_sample.append(data_matriz[current_index])
        current_index = (current_index + K) % N

for r_idx, row_data in enumerate(ms_sample, start=2):
    for c_idx, val in enumerate(row_data):
        sheet_ms.cell(row=r_idx, column=c_idx+1, value=val)

wb.save(output_file)
print(f"File saved successfully to {output_file}")
